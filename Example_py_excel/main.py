import sys
from PyQt5 import uic
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QFileDialog
import openpyxl
from openpyxl import Workbook

from_class = uic.loadUiType("MainDialog.ui")[0]

class MainApp(QMainWindow, from_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.pushButton_import.clicked.connect(self.slot_pushButton_import)  # import 버튼 클릭
        self.pushButton_export.clicked.connect(self.slot_pushButton_export)  # export 버튼 클릭


    def slot_pushButton_export(self):
        print("slot_pushButton_export")
        # 파일 선택
        path = QFileDialog.getSaveFileName(self, 'Save File', '_saveFile', 'All File(*);; xlsx File(*.xlsx)')
        # filePath에 현재 읽어온 엑셀 파일 경로를 입력한다.(절대경로)

        if not path[0]:
            return

        print(path[0])

        # 워크북 생성
        wb = openpyxl.Workbook()

        # 시트 활성화
        ws = wb.worksheets[0]

        # 시트 이름바꾸기
        ws.title = 'test1234'

        count = self.tableWidget.rowCount()
        #   시트에 쓰기
        for row in range(count):
            _row = row + 1
            item_0 = self.tableWidget.item(row, 0)  # A
            _cell = 'A' + str(_row)
            # print(_cell + " text: " + item_0.text())
            ws[_cell] = item_0.text()

            item_1 = self.tableWidget.item(row, 1)  # B
            _cell = 'B' + str(_row)
            # print(_cell + " text: " + item_1.text())
            ws[_cell] = item_1.text()

            item_2 = self.tableWidget.item(row, 2)  # C
            _cell = 'C' + str(_row)
            # print(_cell + " text: " + item_2.text())
            ws[_cell] = item_2.text()

            item_3 = self.tableWidget.item(row, 3)  # D
            _cell = 'D' + str(_row)
            # print(_cell + " text: " + item_3.text())
            ws[_cell] = item_3.text()

            item_4 = self.tableWidget.item(row, 4)  # E
            _cell = 'E' + str(_row)
            # print(_cell + " text: " + item_4.text())
            ws[_cell] = item_4.text()

        # 워크북 filename 으로 저장(없으면 새로생성)
        filename = path[0]
        wb.save(filename)

    def slot_pushButton_import(self):
        print("slot_pushButton_import")

        # 파일 선택
        path = QFileDialog.getOpenFileName(self, 'Open File', '', 'All File(*);; xlsx File(*.xlsx)')
        # filePath에 현재 읽어온 엑셀 파일 경로를 입력한다.(절대경로)

        if not path[0]:
            return

        print(path[0])

        # 위 절대 경로 활용해 openpyxl workbook 객체 생성
        wb = openpyxl.load_workbook(path[0])
        # wb.get_sheet_names()
        # 설정한 workbook의 시트리스트를 읽어온다.
        self.shtlist = wb.sheetnames
        print(self.shtlist)

        #sheet = wb['test']
        sheet = wb[self.shtlist[0]]

        print("slot_pushButton_import11")

        # multiple_cells = sheet['A':'K']
        # for row in multiple_cells:
        #    for cell in row:
        #        print (cell.value)

        while self.tableWidget.rowCount() > 0:  # 테이블 위젯 모두 지우기
            self.tableWidget.removeRow(0)

        for col in sheet['A']:
            print(col.value)
            _value = col.value
            # row item 추가
            _rowCount = self.tableWidget.rowCount()
            self.tableWidget.insertRow(_rowCount);
            self.tableWidget.setItem(_rowCount, 0, QTableWidgetItem(str(_value)))

        _temp = 0
        for col in sheet['B']:
            print(col.value)
            # row item 추가
            self.tableWidget.setItem(_temp, 1, QTableWidgetItem(col.value))
            _temp += 1

        _temp = 0
        for col in sheet['C']:
            print(col.value)
            # row item 추가
            self.tableWidget.setItem(_temp, 2, QTableWidgetItem(col.value))
            _temp += 1

        _temp = 0
        for col in sheet['D']:
            print(col.value)
            # row item 추가
            self.tableWidget.setItem(_temp, 3, QTableWidgetItem(col.value))
            _temp += 1

        _temp = 0
        for col in sheet['E']:
            print(col.value)
            # row item 추가
            self.tableWidget.setItem(_temp, 4, QTableWidgetItem(col.value))
            _temp += 1



if __name__ == '__main__':
    app = QApplication(sys.argv)
    myWindow = MainApp()
    myWindow.show()
    app.exec_()
